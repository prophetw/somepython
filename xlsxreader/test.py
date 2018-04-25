#coding:UTF-8
from openpyxl import load_workbook
from pprint import pprint
import json


# 脚本从 .xlsx 生成类似 json
# 脚本依赖 openpyxl  需要使用 pip install openpyxl 安装
# openpyxl docs https://openpyxl.readthedocs.io/en/stable/tutorial.html#loading-from-a-file
# 尽量 python3+ 版本

#  需要设置4个参数 如下所示
#       path 是xlsx的路径
#       sheetName 表格的名字 注意: 这个sheetName请手动改成英文 (python3+可以不用考虑修改,python2+对中文支持不好会报错)
#       keyColumn key列
#       valueColumn 结果列
#  另外 keyMap 对象需要手动维护 保持提供的xlsx里面的key和翻译文件的key一一对应

path = './1.xlsx'
sheetName = str('Page title And Meta Description')
keyColumn = 'B'   # xlsx B列
valueColumn = 'E'  # xlsx D列
print('xlsx path:',path)
print('sheetName:',sheetName)
print('keyColumn:',keyColumn)
print('valueColumn',valueColumn)

# 关于xlsx的格式
# https://jira.foxitsoftware.cn/browse/PRO-5820 需要PM 严格按照里面的xlsx格式编写 英文和中文同一行

# 实际效果如下图所示
# input xlsx
#  A         B列                C               D列                      E
#  1         key                                page title              meta description
#  2         PDF to Word                        hello title\n你好标题    hello meta description\n你好描述
#
#  当设置 keyColumn='B' valueColumn = 'D'
#
# output json
# {
#     "PDF2Word":"hello"
# }
# {
#     "PDF2Word":"你好标题"
# }


wb = load_workbook(filename=path)
ws = wb[sheetName]

# 提供xlsx里面的key 和 翻译文件的key
keyMap={
    "PDF to Word" : "PDF2Word",
    "PDF to Excel":"PDF2Excel",
    "PDF to PPT":"PDF2PPT",
    "PDF to Image":"PDF2Image",
    "PDF to Text":"PDF2Text",
    "PDF to HTML":"PDF2HTML",
    "Word  to PDF":"Word2PDF",
    "Excel  to PDF":"Excel2PDF",
    "PPT to PDF":"PPT2PDF",
    "Image to PDF":"Image2PDF",
    "HTML to PDF":"HTML2PDF",
    "Text to PDF":"Text2PDF",
    "Merge PDF":"MergePDF",
    "Watermark PDF":"Watermark",
    "Compress PDF": "CompressPDF",
    "Split PDF":"SplitPDF",
    "PDF Organizer":"PageOrganizer",
    "Redact PDF":"RedactPDF",
    "Flatten PDF":"FlattenPDF",
    "PDF Header Footer":"HeaderFooter",
    "Password protect PDF":"PasswordProtect",
    "PDF to cPDF":"PDF2CPDF",
    "Foxit Drive":"Foxit_Drive",
    "phantom pdf online":"PhantomPDF_Online",
    "Foxit Online":"Foxit_Online"
}

# 遍历30行足够了
rowNumber = range(1, 30)

# 输出对象
outputEn = {}
outputZh = {}

for index in rowNumber:
    if ws[keyColumn+str(index)].value != None and keyMap.get(ws[keyColumn+str(index)].value) != None:
        key = keyMap[ws[keyColumn+str(index)].value]
        outputEn[key] = ws[valueColumn + str(index)].value.split('\n')[0]
        outputZh[key] = ws[valueColumn + str(index)].value.split('\n')[1]

# python 3 如下
pprint(outputEn,width=999)
pprint(outputZh,width=999)
print('输出格式不对的时候使用下面这个')
pprint(outputZh)

# python 2 版本中文字典打印出 unicode 的问题用下面的方案
# print(json.dumps(metaDescriptionZh, ensure_ascii=False, encoding='UTF-8'))
